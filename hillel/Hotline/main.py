import csv
import random
import time
import sqlite3
from typing import Tuple, Dict, List

import openpyxl
from bs4 import BeautifulSoup

import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

class UserAgent:
    headers = (
        "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.134 Safari/537.36 OPR/89.0.4447.71",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3679.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) KnoxTeams/2.1.32 Chrome/102.0.5005.167 Electron/19.0.12 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.1758.95",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.1961.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 OPR/109.0.0.0"
    )

    def get_random_agent(self) -> str:
        idx = random.randint(0, len(self.headers) - 1)
        return self.headers[idx]

    def random(self) -> dict:
        agent = self.get_random_agent()
        accept = 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
        return {'User-Agent': agent, 'Accept': accept}


class BaseItem:
    fields: Tuple

    def to_dict(self) -> Dict:
        return {field: getattr(self, field) for field in self.fields}

    def to_tuple(self) -> Tuple:
        return tuple(getattr(self, field) for field in self.fields)


class MoBile(BaseItem):

    def __init__(self, title, href, price):
        self.title: str = title
        self.href: str = href
        self.price: str = price
        self.fields: tuple = tuple(self.__dict__.keys())

class Request:

    def __init__(self):
        self.service = None
        self.driver = None
        self.data = None

    def set_selenium_settings(self, executable_path):
        self.service = Service(executable_path=executable_path)

    # def get_data_by_selenium(self, url) -> str:
    #     driver = webdriver.Edge(service=self.service)
    #     driver.get(url)
    #     time.sleep(3)
    #     self.data = driver.page_source
    #     driver.close()
    #     driver.quit()
    #     return self.data

    def get_data_by_requests(self, url) -> str:
        header = UserAgent().random()
        resp = requests.get(url, headers=header)
        if resp.status_code == 200:
            self.data = resp.text

        return self.data

    def save_to_file(self, name='draft.html'):
        with open(name, 'w', encoding='utf8') as file:
            file.write(self.data)

class Parser:

    def __init__(self, item_class):
        self.rows: List = []
        self.item_class = item_class

    def parse_data(self, url, func):
        request = Request()
        # for i in range(1, 3):
        # url = url.format(i)
        html = request.get_data_by_requests(url)
        if html:
            soup = BeautifulSoup(html, 'html.parser')
            self.rows += func(soup, self.item_class)
                # time.sleep(2)

    # def save_to_csv(self, name='mobile.csv'):
    #     if self.rows:
    #         rows = [card.to_dict() for card in self.rows]
    #         row = rows[0]
    #         csv_title = list(row.keys())
    #         with open(name, 'w') as f:
    #             writer = csv.DictWriter(f, fieldnames=csv_title, delimiter=';')
    #             writer.writeheader()
    #             writer.writerows(rows)

    def save_to_excel(self, name='mobile.xlsx'):
        rows = [card.to_dict() for card in self.rows]
        if rows:
            wb = openpyxl.Workbook()
            ws = wb.active
            row = rows[0]
            titles = list(row.keys())
            row_num = 1
            for i, key in enumerate(titles, 1):
                ws.cell(row=row_num, column=i).value = key
            row_num += 1
            for row in rows:
                for i, val in enumerate(row.values(), 1):
                    ws.cell(row=row_num, column=i).value = val
                row_num += 1
            wb.save(name)

    def save_to_db(self, data_base, table_name):
        rows = [card.to_tuple() for card in self.rows]
        data_base.save(table_name, rows)


class DataBase:

    def __init__(self, db_name):
        self.db_name: str = db_name

    def create_table(self, table_name):
        sqlite_connection = sqlite3.connect(self.db_name)
        sqlite_create_table_query = f'''
                CREATE TABLE IF NOT EXISTS {table_name} (
                    id INTEGER PRIMARY KEY,
                    title TEXT NOT NULL,
                    url TEXT NOT NULL,
                    price TEXT NOT NULL           
                );'''
        cursor = sqlite_connection.cursor()
        cursor.execute(sqlite_create_table_query)
        sqlite_connection.commit()
        print("Created")

    def save(self, table_name, rows):
        sqlite_connection = sqlite3.connect(self.db_name)
        cursor = sqlite_connection.cursor()
        cursor.executemany(
            f"""INSERT INTO {table_name} 
            ('title', 'url', 'price' ) VALUES (?,?,?)""",
            rows)
        sqlite_connection.commit()
        print("Saved")

def parse_hotline_mob_cards(soup, item_class) -> List:
    res = []
    li_list = soup.find_all('div', attrs={'class': 'list-item'})
    for li in li_list:
        a = li.find('div', attrs={'class': 'list-item__title-container'})
        if a:
            title = a.text
            title = title.strip()
            a_tag = a.find('a')
            if a_tag:
                d = 'https://hotline.ua'
                href1 = a_tag['href']
                href = d + href1
            # href = a.get('href')
            # href = a['href']
            # old_price = ''
            # old_price_raw = li.find('div', attrs={'old-price'})
            price_raw = li.find('div', attrs={'class': 'list-item__value-price'})
            # if old_price_raw:
            #     old_text = old_price_raw.text
            #     if old_text:
            #         old_price = int(''.join(c for c in old_text if c.isdigit()))
            price = ''.join(c for c in price_raw.text if c.isdigit() or c == '–')
            b = 1
            res.append(
                item_class(title=title, href=href,
                           price=price)
            )
    return res


if __name__ == '__main__':
    url = 'https://hotline.ua/ua/mobile/mobilnye-telefony-i-smartfony/'
    # request = Request()
    # html = request.get_data_by_requests(url)
    # request.save_to_file()
    parser = Parser(MoBile)
    parser.parse_data(url, parse_hotline_mob_cards)
    # parser.save_to_excel(name='mobile.xlsx')
    data_base = DataBase('mob.db')
    data_base.create_table('mobiles')
    parser.save_to_db(data_base, 'mobiles')